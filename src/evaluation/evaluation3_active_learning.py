# -*- coding: utf-8 -*-
"""Active-learning utilities for EVALUATION3 duplicate-pair classification.

This module operates on the pair-level evidence emitted by NB10D
(`evaluation3_manual_review_KEY.csv`).  It does NOT regenerate images or run the
full overlap audit.  The intended workflow is:

1. Reserve fixed validation and final-test pairs (group-disjoint by E3 outfit).
2. Human-label a small diverse seed set.
3. Fit several classical ML models on the accumulated training labels.
4. Query the most uncertain remaining pairs for the next human-review round.
5. Refit after each round.
6. Select triage thresholds only on validation data.
7. Touch the final test exactly once after model + thresholds are frozen.

The classifier target is binary:
    DUPLICATE -> 1
    NON_DUPLICATE -> 0

The middle probability region is intentionally left for MANUAL_REVIEW.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import numpy as np
import pandas as pd

from sklearn.cluster import KMeans
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    average_precision_score,
    balanced_accuracy_score,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC


DUPLICATE = "DUPLICATE"
NON_DUPLICATE = "NON_DUPLICATE"
VALID_LABELS = frozenset({DUPLICATE, NON_DUPLICATE})

# Raw evidence columns that exist in NB10D manual KEY rows.
RAW_FEATURE_COLUMNS = (
    "phash_distance",
    "alignment_success",
    "ecc_correlation",
    "rotation_degrees",
    "translation_x",
    "translation_y",
    "rgb_ssim",
    "gray_ssim",
    "foreground_iou",
    "edge_ssim",
    "mean_lab_delta",
    "interior_mae",
    "patch_mae_max",
    "patch_mae_p90",
    "patch_count",
)

# Stable ML feature names after deterministic transforms.
MODEL_FEATURE_COLUMNS = (
    "phash_distance",
    "alignment_success",
    "ecc_correlation",
    "abs_rotation_degrees",
    "abs_translation_x",
    "abs_translation_y",
    "rgb_ssim",
    "gray_ssim",
    "foreground_iou",
    "edge_ssim",
    "mean_lab_delta",
    "interior_mae",
    "patch_mae_max",
    "patch_mae_p90",
    "patch_count",
)


@dataclass(frozen=True)
class TriageThresholds:
    auto_non_max_probability: float
    auto_duplicate_min_probability: float
    auto_non_npv: float
    auto_duplicate_precision: float
    auto_non_count: int
    auto_duplicate_count: int
    manual_count: int


def _as_bool01(series: pd.Series) -> pd.Series:
    if series.dtype == bool:
        return series.astype(float)
    normalized = series.astype(str).str.strip().str.lower()
    return normalized.map(
        {
            "true": 1.0,
            "1": 1.0,
            "yes": 1.0,
            "false": 0.0,
            "0": 0.0,
            "no": 0.0,
        }
    ).fillna(0.0)


def prepare_feature_frame(rows: pd.DataFrame) -> pd.DataFrame:
    """Return numeric model features with deterministic nuisance transforms."""

    missing = [column for column in RAW_FEATURE_COLUMNS if column not in rows.columns]
    if missing:
        raise KeyError(f"Missing NB10D evidence columns: {missing}")

    result = pd.DataFrame(index=rows.index)
    result["phash_distance"] = pd.to_numeric(rows["phash_distance"], errors="coerce")
    result["alignment_success"] = _as_bool01(rows["alignment_success"])
    result["ecc_correlation"] = pd.to_numeric(rows["ecc_correlation"], errors="coerce")
    result["abs_rotation_degrees"] = pd.to_numeric(
        rows["rotation_degrees"], errors="coerce"
    ).abs()
    result["abs_translation_x"] = pd.to_numeric(
        rows["translation_x"], errors="coerce"
    ).abs()
    result["abs_translation_y"] = pd.to_numeric(
        rows["translation_y"], errors="coerce"
    ).abs()

    for column in (
        "rgb_ssim",
        "gray_ssim",
        "foreground_iou",
        "edge_ssim",
        "mean_lab_delta",
        "interior_mae",
        "patch_mae_max",
        "patch_mae_p90",
        "patch_count",
    ):
        result[column] = pd.to_numeric(rows[column], errors="coerce")

    # Classical sklearn estimators used here do not accept NaN.  Median filling
    # is intentionally simple/auditable for this pilot.
    for column in result.columns:
        values = result[column]
        median = float(values.median()) if values.notna().any() else 0.0
        result[column] = values.fillna(median).astype(float)

    return result.loc[:, list(MODEL_FEATURE_COLUMNS)]


def label_to_target(label: object) -> int | None:
    value = str(label or "").strip().upper()
    if value == DUPLICATE:
        return 1
    if value == NON_DUPLICATE:
        return 0
    return None


def read_review_file(path: Path | str) -> pd.DataFrame:
    """Read a review CSV/XLSX and normalize pair_id/human_label."""

    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(source)
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(source)
    else:
        frame = pd.read_csv(source)
    if "pair_id" not in frame.columns or "human_label" not in frame.columns:
        raise KeyError(f"{source} must contain pair_id and human_label")
    frame = frame.copy()
    frame["pair_id"] = frame["pair_id"].astype(str).str.strip()
    frame["human_label"] = frame["human_label"].astype(str).str.strip().str.upper()
    frame = frame[frame["human_label"].isin(VALID_LABELS)]
    return frame


def merge_review_labels(
    key_rows: pd.DataFrame,
    review_paths: Iterable[Path | str],
) -> pd.DataFrame:
    """Merge all valid human labels onto NB10D key rows; latest file wins."""

    labels: dict[str, str] = {}
    for path in review_paths:
        source = Path(path)
        if not source.is_file():
            continue
        review = read_review_file(source)
        for row in review.itertuples(index=False):
            labels[str(row.pair_id)] = str(row.human_label)

    result = key_rows.copy()
    result["pair_id"] = result["pair_id"].astype(str).str.strip()
    result["human_label"] = result["pair_id"].map(labels).fillna("")
    result["target"] = result["human_label"].map(label_to_target)
    return result


def _candidate_one_per_group(
    frame: pd.DataFrame,
    *,
    group_column: str,
    random_state: int,
) -> pd.DataFrame:
    if group_column not in frame.columns:
        return frame.copy()
    shuffled = frame.sample(frac=1.0, random_state=random_state)
    return shuffled.drop_duplicates(subset=[group_column], keep="first")


def diverse_batch(
    frame: pd.DataFrame,
    *,
    batch_size: int,
    group_column: str = "e3_outfit_id",
    random_state: int = 42,
) -> pd.DataFrame:
    """Select a feature-diverse, roughly one-per-outfit review batch."""

    if frame.empty or batch_size <= 0:
        return frame.iloc[0:0].copy()
    candidates = _candidate_one_per_group(
        frame,
        group_column=group_column,
        random_state=random_state,
    )
    if len(candidates) <= batch_size:
        return candidates.copy()

    features = prepare_feature_frame(candidates)
    scaled = StandardScaler().fit_transform(features)
    cluster_count = min(int(batch_size), len(candidates))
    kmeans = KMeans(
        n_clusters=cluster_count,
        n_init=10,
        random_state=random_state,
    )
    cluster_ids = kmeans.fit_predict(scaled)
    centers = kmeans.cluster_centers_

    selected_positions: list[int] = []
    for cluster_id in range(cluster_count):
        positions = np.flatnonzero(cluster_ids == cluster_id)
        local = scaled[positions]
        distances = np.linalg.norm(local - centers[cluster_id], axis=1)
        selected_positions.append(int(positions[int(np.argmin(distances))]))

    return candidates.iloc[selected_positions].copy()


def build_models(random_state: int = 42) -> dict[str, object]:
    """Return the three intentionally small baseline model families."""

    return {
        "logistic_regression": Pipeline(
            [
                ("scale", StandardScaler()),
                (
                    "model",
                    LogisticRegression(
                        C=1.0,
                        class_weight="balanced",
                        max_iter=3000,
                        random_state=random_state,
                    ),
                ),
            ]
        ),
        "svm_rbf": Pipeline(
            [
                ("scale", StandardScaler()),
                (
                    "model",
                    SVC(
                        C=2.0,
                        kernel="rbf",
                        gamma="scale",
                        class_weight="balanced",
                        probability=True,
                        random_state=random_state,
                    ),
                ),
            ]
        ),
        "random_forest": RandomForestClassifier(
            n_estimators=400,
            min_samples_leaf=2,
            max_features="sqrt",
            class_weight="balanced_subsample",
            n_jobs=-1,
            random_state=random_state,
        ),
    }


def _positive_probability(model, features: pd.DataFrame) -> np.ndarray:
    probabilities = model.predict_proba(features)
    classes = list(model.classes_)
    positive_index = classes.index(1)
    return np.asarray(probabilities[:, positive_index], dtype=float)


def binary_metrics(y_true: Sequence[int], probability: Sequence[float]) -> dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(probability, dtype=float)
    prediction = (p >= 0.5).astype(int)
    metrics = {
        "balanced_accuracy@0.5": float(balanced_accuracy_score(y, prediction)),
        "precision@0.5": float(precision_score(y, prediction, zero_division=0)),
        "recall@0.5": float(recall_score(y, prediction, zero_division=0)),
        "f1@0.5": float(f1_score(y, prediction, zero_division=0)),
    }
    if len(np.unique(y)) == 2:
        metrics["roc_auc"] = float(roc_auc_score(y, p))
        metrics["pr_auc"] = float(average_precision_score(y, p))
    else:
        metrics["roc_auc"] = float("nan")
        metrics["pr_auc"] = float("nan")
    return metrics


def fit_compare_models(
    train_rows: pd.DataFrame,
    validation_rows: pd.DataFrame,
    *,
    random_state: int = 42,
) -> tuple[dict[str, object], pd.DataFrame]:
    """Fit LR/RBF-SVM/RF and score each on one fixed validation set."""

    if train_rows["target"].isna().any() or validation_rows["target"].isna().any():
        raise ValueError("train/validation rows must all have human labels")
    if train_rows["target"].nunique() < 2:
        raise ValueError("training labels need both DUPLICATE and NON_DUPLICATE")

    x_train = prepare_feature_frame(train_rows)
    y_train = train_rows["target"].astype(int).to_numpy()
    x_val = prepare_feature_frame(validation_rows)
    y_val = validation_rows["target"].astype(int).to_numpy()

    fitted: dict[str, object] = {}
    records: list[dict[str, object]] = []
    for name, model in build_models(random_state=random_state).items():
        model.fit(x_train, y_train)
        probability = _positive_probability(model, x_val)
        record = {"model": name, **binary_metrics(y_val, probability)}
        fitted[name] = model
        records.append(record)

    report = pd.DataFrame(records)
    # AUC is primary for model ranking here; balanced accuracy breaks ties.
    report = report.sort_values(
        by=["roc_auc", "balanced_accuracy@0.5", "pr_auc"],
        ascending=False,
        na_position="last",
    ).reset_index(drop=True)
    return fitted, report


def uncertainty_batch(
    model,
    pool_rows: pd.DataFrame,
    *,
    batch_size: int,
    group_column: str = "e3_outfit_id",
) -> pd.DataFrame:
    """Query pairs closest to p=0.5, while avoiding repeated outfits first."""

    if pool_rows.empty or batch_size <= 0:
        return pool_rows.iloc[0:0].copy()
    features = prepare_feature_frame(pool_rows)
    probability = _positive_probability(model, features)
    scored = pool_rows.copy()
    scored["model_probability_duplicate"] = probability
    scored["uncertainty"] = np.abs(probability - 0.5)
    scored = scored.sort_values(
        ["uncertainty", "model_probability_duplicate"],
        ascending=[True, True],
    )

    selected_indices: list[object] = []
    seen_groups: set[str] = set()
    for index, row in scored.iterrows():
        group = str(row.get(group_column, ""))
        if group and group in seen_groups:
            continue
        selected_indices.append(index)
        if group:
            seen_groups.add(group)
        if len(selected_indices) >= batch_size:
            break

    if len(selected_indices) < batch_size:
        for index in scored.index:
            if index in selected_indices:
                continue
            selected_indices.append(index)
            if len(selected_indices) >= batch_size:
                break

    return scored.loc[selected_indices].copy()


def choose_triage_thresholds(
    y_true: Sequence[int],
    probability_duplicate: Sequence[float],
    *,
    target_auto_duplicate_precision: float = 0.98,
    target_auto_non_npv: float = 0.995,
    minimum_auto_examples: int = 5,
) -> TriageThresholds:
    """Choose two validation-only thresholds for DUP / MANUAL / NON triage.

    - p >= high -> AUTO DUPLICATE, selected for high positive precision.
    - p <= low  -> AUTO NON_DUPLICATE, selected for high NPV (clean precision).
    - otherwise -> MANUAL_REVIEW.

    If the validation sample cannot support a requested guarantee, the
    corresponding automatic region collapses (high=1.0 or low=0.0).
    """

    y = np.asarray(y_true, dtype=int)
    p = np.asarray(probability_duplicate, dtype=float)
    if y.shape != p.shape:
        raise ValueError("y_true and probability_duplicate must have same shape")
    if len(y) == 0:
        raise ValueError("validation set is empty")

    high = 1.0
    high_precision = 1.0
    high_count = 0
    for threshold in sorted(np.unique(p)):
        mask = p >= float(threshold)
        count = int(mask.sum())
        if count < int(minimum_auto_examples):
            continue
        precision = float(y[mask].mean())
        if precision >= float(target_auto_duplicate_precision):
            high = float(threshold)
            high_precision = precision
            high_count = count
            break

    low = 0.0
    low_npv = 1.0
    low_count = 0
    for threshold in sorted(np.unique(p), reverse=True):
        mask = p <= float(threshold)
        count = int(mask.sum())
        if count < int(minimum_auto_examples):
            continue
        # NPV = fraction truly NON among examples auto-declared NON.
        npv = float((1 - y[mask]).mean())
        if npv >= float(target_auto_non_npv):
            low = float(threshold)
            low_npv = npv
            low_count = count
            break

    if low >= high:
        # No safe separation on this validation set.  Prefer manual review.
        low, high = 0.0, 1.0
        low_npv, high_precision = 1.0, 1.0
        low_count, high_count = 0, 0

    manual_count = int(((p > low) & (p < high)).sum())
    return TriageThresholds(
        auto_non_max_probability=low,
        auto_duplicate_min_probability=high,
        auto_non_npv=low_npv,
        auto_duplicate_precision=high_precision,
        auto_non_count=low_count,
        auto_duplicate_count=high_count,
        manual_count=manual_count,
    )


def apply_triage(
    probability_duplicate: Sequence[float],
    thresholds: TriageThresholds,
) -> np.ndarray:
    p = np.asarray(probability_duplicate, dtype=float)
    result = np.full(p.shape, "MANUAL_REVIEW", dtype=object)
    result[p <= thresholds.auto_non_max_probability] = NON_DUPLICATE
    result[p >= thresholds.auto_duplicate_min_probability] = DUPLICATE
    return result


def write_review_workbook(
    rows: pd.DataFrame,
    destination: Path | str,
    *,
    title: str,
) -> Path:
    """Write a compact human-label workbook with preview paths and evidence."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(destination)
    path.parent.mkdir(parents=True, exist_ok=True)

    display_columns = [
        "pair_id",
        "preview_file",
        "e3_outfit_id",
        "e3_slot",
        "phash_distance",
        "ecc_correlation",
        "rgb_ssim",
        "edge_ssim",
        "mean_lab_delta",
        "interior_mae",
        "patch_mae_max",
        "model_probability_duplicate",
        "human_label",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(display_columns)
    for _, row in rows.iterrows():
        values = []
        for column in display_columns:
            if column == "human_label":
                values.append("")
            else:
                value = row.get(column, "")
                if pd.isna(value):
                    value = ""
                values.append(value)
        sheet.append(values)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    widths = {
        "A": 16,
        "B": 48,
        "C": 18,
        "D": 10,
        "E": 14,
        "F": 16,
        "G": 14,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 14,
        "L": 20,
        "M": 20,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    validation = DataValidation(
        type="list",
        formula1='"DUPLICATE,NON_DUPLICATE"',
        allow_blank=True,
    )
    validation.error = "Chỉ chọn DUPLICATE hoặc NON_DUPLICATE."
    validation.errorTitle = "Invalid label"
    sheet.add_data_validation(validation)
    if len(rows):
        validation.add(f"M2:M{len(rows) + 1}")

    sheet["A1"].comment = None
    sheet.sheet_properties.tabColor = "1F4E78"
    workbook.properties.title = title
    workbook.save(path)
    return path

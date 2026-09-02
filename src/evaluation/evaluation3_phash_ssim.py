# -*- coding: utf-8 -*-
"""EVALUATION3 ↔ Polyvore overlap audit using calibrated pHash + SSIM.

This module is intentionally separate from the original dHash NB10 path.
The canonical decision states are only:

- DUPLICATE
- MANUAL_REVIEW
- NON_DUPLICATE

Frozen calibrated rule (same-image overlap):

- pHash distance <= 4 and SSIM >= 0.92 -> DUPLICATE
- pHash distance <= 4 and 0.90 <= SSIM < 0.92 -> MANUAL_REVIEW
- everything else -> NON_DUPLICATE

An optional exact-pixel shortcut can be enabled as an ablation. When enabled,
a normalized decoded-pixel SHA256 match is immediately DUPLICATE. ID collisions
are not used in the decision path.
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, MutableMapping, Sequence

import numpy as np
from scipy.fftpack import dct
from skimage.metrics import structural_similarity

from src.evaluation.evaluation3_overlap import (
    BKTree,
    DEFAULT_MODEL_DEVELOPMENT_SPLITS,
    IMAGE_SUFFIXES,
    _candidate_item_ids,
    _require_pillow,
    discover_evaluation3_outfits,
    load_development_identity,
    load_evaluation3_annotations,
    merge_evaluation3_annotations,
    write_json,
    write_jsonl,
)


DUPLICATE = "DUPLICATE"
MANUAL_REVIEW = "MANUAL_REVIEW"
NON_DUPLICATE = "NON_DUPLICATE"
VALID_FINAL_MANUAL_LABELS = frozenset({DUPLICATE, NON_DUPLICATE})

DEFAULT_PHASH_HAMMING_THRESHOLD = 4
DEFAULT_SSIM_AUTO_DUPLICATE_THRESHOLD = 0.92
DEFAULT_SSIM_MANUAL_LOWER_BOUND = 0.90
DEFAULT_SSIM_SIZE = 256
PHASH_HASH_SIZE = 8
PHASH_HIGHFREQ_FACTOR = 4
FINGERPRINT_VERSION = "phash64-dct-v1_rgb-pixel-sha256-v1"
MAX_MATCH_EXAMPLES_PER_IMAGE = 20


@dataclass(frozen=True)
class PerceptualFingerprint:
    phash64: int
    pixel_sha256: str
    width: int
    height: int


@dataclass(frozen=True)
class PerceptualDevelopmentItem:
    item_id: str
    splits: tuple[str, ...]
    source_kit_ids: tuple[str, ...]
    fingerprint: PerceptualFingerprint
    image_source: str


class ImageStore:
    """Load Polyvore images on demand and cache SSIM-ready grayscale arrays."""

    def __init__(
        self,
        *,
        ssim_size: int = DEFAULT_SSIM_SIZE,
        cache_size: int = 1024,
    ) -> None:
        self.ssim_size = int(ssim_size)
        self.cache_size = int(cache_size)
        self._ssim_cache: OrderedDict[str, np.ndarray] = OrderedDict()

    def load_image(self, item_id: str):  # pragma: no cover - abstract contract
        raise NotImplementedError

    def ssim_array(self, item_id: str) -> np.ndarray:
        cached = self._ssim_cache.get(item_id)
        if cached is not None:
            self._ssim_cache.move_to_end(item_id)
            return cached
        image = self.load_image(item_id)
        array = image_to_ssim_array(image, size=self.ssim_size)
        self._ssim_cache[item_id] = array
        if len(self._ssim_cache) > self.cache_size:
            self._ssim_cache.popitem(last=False)
        return array


class LocalImageStore(ImageStore):
    def __init__(self, paths: Mapping[str, Path], **kwargs) -> None:
        super().__init__(**kwargs)
        self.paths = dict(paths)

    def load_image(self, item_id: str):
        Image, _ = _require_pillow()
        path = self.paths[item_id]
        with Image.open(path) as image:
            image.load()
            return image.copy()


class HuggingFaceImageStore(ImageStore):
    def __init__(
        self,
        datasets_by_split: Mapping[str, object],
        locators: Mapping[str, tuple[str, int]],
        **kwargs,
    ) -> None:
        super().__init__(**kwargs)
        self.datasets_by_split = dict(datasets_by_split)
        self.locators = dict(locators)

    def load_image(self, item_id: str):
        Image, _ = _require_pillow()
        split, row_index = self.locators[item_id]
        value = self.datasets_by_split[split][row_index]["image"]
        if isinstance(value, Image.Image):
            return value.copy()
        if isinstance(value, Mapping):
            if value.get("bytes") is not None:
                from io import BytesIO

                with Image.open(BytesIO(value["bytes"])) as image:
                    image.load()
                    return image.copy()
            if value.get("path"):
                with Image.open(value["path"]) as image:
                    image.load()
                    return image.copy()
        if isinstance(value, (str, Path)):
            with Image.open(value) as image:
                image.load()
                return image.copy()
        raise ValueError(
            f"Unsupported Hugging Face image value: {type(value).__name__}"
        )


def normalized_pixel_sha256(image) -> tuple[str, int, int]:
    """Hash EXIF-normalized RGB decoded pixels, including dimensions."""

    _, ImageOps = _require_pillow()
    normalized = ImageOps.exif_transpose(image).convert("RGB")
    width, height = normalized.size
    digest = hashlib.sha256()
    digest.update(f"RGB:{width}x{height}:".encode("ascii"))
    digest.update(normalized.tobytes())
    return digest.hexdigest(), width, height


def phash64_image(image) -> int:
    """Standard 64-bit DCT pHash compatible with the common ImageHash recipe."""

    Image, ImageOps = _require_pillow()
    normalized = ImageOps.exif_transpose(image).convert("L")
    size = PHASH_HASH_SIZE * PHASH_HIGHFREQ_FACTOR
    resampling = getattr(Image, "Resampling", Image)
    normalized = normalized.resize((size, size), resampling.LANCZOS)
    pixels = np.asarray(normalized, dtype=np.float32)
    transformed = dct(dct(pixels, axis=0), axis=1)
    low_frequency = transformed[:PHASH_HASH_SIZE, :PHASH_HASH_SIZE]
    median = float(np.median(low_frequency))
    bits = low_frequency > median
    value = 0
    for bit in bits.reshape(-1):
        value = (value << 1) | int(bool(bit))
    return value


def fingerprint_image(image) -> PerceptualFingerprint:
    pixel_sha256, width, height = normalized_pixel_sha256(image)
    return PerceptualFingerprint(
        phash64=phash64_image(image),
        pixel_sha256=pixel_sha256,
        width=width,
        height=height,
    )


def fingerprint_path(path: Path | str) -> PerceptualFingerprint:
    Image, _ = _require_pillow()
    source = Path(path)
    try:
        with Image.open(source) as image:
            image.load()
            return fingerprint_image(image)
    except Exception as error:
        raise ValueError(f"Cannot decode image: {source}: {error}") from error


def image_to_ssim_array(image, *, size: int = DEFAULT_SSIM_SIZE) -> np.ndarray:
    """Normalize image for SSIM: EXIF transpose -> grayscale -> square resize."""

    Image, ImageOps = _require_pillow()
    normalized = ImageOps.exif_transpose(image).convert("L")
    resampling = getattr(Image, "Resampling", Image)
    normalized = normalized.resize((int(size), int(size)), resampling.LANCZOS)
    return np.asarray(normalized, dtype=np.uint8)


def ssim_score(left: np.ndarray, right: np.ndarray) -> float:
    if left.shape != right.shape:
        raise ValueError(
            f"SSIM arrays must have the same shape: {left.shape} vs {right.shape}"
        )
    return float(structural_similarity(left, right, data_range=255))


def hamming_distance(left: int, right: int) -> int:
    return (left ^ right).bit_count()


def classify_pair(
    *,
    phash_distance: int,
    ssim: float | None,
    exact_pixel_match: bool = False,
    use_exact_pixel: bool = False,
    phash_threshold: int = DEFAULT_PHASH_HAMMING_THRESHOLD,
    ssim_auto_threshold: float = DEFAULT_SSIM_AUTO_DUPLICATE_THRESHOLD,
    ssim_manual_lower_bound: float = DEFAULT_SSIM_MANUAL_LOWER_BOUND,
) -> tuple[str, str]:
    """Return (decision, method) for one E3↔Polyvore image pair."""

    if use_exact_pixel and exact_pixel_match:
        return DUPLICATE, "exact_pixel"
    if phash_distance > phash_threshold:
        return NON_DUPLICATE, "phash_outside_radius"
    if ssim is None:
        raise ValueError(
            "SSIM is required when pHash distance is inside the candidate radius"
        )
    if ssim >= ssim_auto_threshold:
        return DUPLICATE, "phash_ssim_auto"
    if ssim >= ssim_manual_lower_bound:
        return MANUAL_REVIEW, "phash_ssim_manual"
    return NON_DUPLICATE, "phash_ssim_non_duplicate"


def _fingerprint_hf_value(value) -> PerceptualFingerprint:
    Image, _ = _require_pillow()
    if isinstance(value, Image.Image):
        return fingerprint_image(value)
    if isinstance(value, Mapping):
        if value.get("bytes") is not None:
            from io import BytesIO

            with Image.open(BytesIO(value["bytes"])) as image:
                image.load()
                return fingerprint_image(image)
        if value.get("path"):
            return fingerprint_path(value["path"])
    if isinstance(value, (str, Path)):
        return fingerprint_path(value)
    raise ValueError(
        f"Unsupported Hugging Face image value: {type(value).__name__}"
    )


def build_local_development_index(
    image_root: Path | str,
    identity: Mapping[str, object],
    *,
    ssim_size: int = DEFAULT_SSIM_SIZE,
) -> tuple[
    list[PerceptualDevelopmentItem], LocalImageStore, dict[str, object], list[str]
]:
    root = Path(image_root).expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"Polyvore image root not found: {root}")

    item_splits = identity["item_splits"]
    item_kits = identity["item_kits"]
    required = set(item_splits)
    paths: dict[str, Path] = {}
    scanned = 0
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_SUFFIXES:
            continue
        scanned += 1
        for item_id in required.intersection(_candidate_item_ids(path, root)):
            paths.setdefault(item_id, path)

    unresolved = sorted(required - set(paths))
    items: list[PerceptualDevelopmentItem] = []
    for item_id in sorted(paths):
        fingerprint = fingerprint_path(paths[item_id])
        items.append(
            PerceptualDevelopmentItem(
                item_id=item_id,
                splits=tuple(sorted(item_splits[item_id])),
                source_kit_ids=tuple(sorted(item_kits[item_id])),
                fingerprint=fingerprint,
                image_source=str(paths[item_id]),
            )
        )
    store = LocalImageStore(paths, ssim_size=ssim_size)
    report = {
        "provider": "local_directory",
        "root": str(root),
        "files_scanned": scanned,
        "resolved_item_count": len(items),
        "fingerprint_version": FINGERPRINT_VERSION,
    }
    return items, store, report, unresolved


def build_huggingface_development_index(
    dataset_name: str,
    split_names: Mapping[str, str],
    identity: Mapping[str, object],
    *,
    ssim_size: int = DEFAULT_SSIM_SIZE,
) -> tuple[
    list[PerceptualDevelopmentItem],
    HuggingFaceImageStore,
    dict[str, object],
    list[str],
]:
    try:
        from datasets import load_dataset
    except ImportError as error:
        raise RuntimeError(
            "datasets is required for the Hugging Face Polyvore provider. "
            "Install requirements-evaluation.txt."
        ) from error

    item_splits = identity["item_splits"]
    item_kits = identity["item_kits"]
    required_by_split: MutableMapping[str, set[str]] = defaultdict(set)
    for item_id, splits in item_splits.items():
        for split in splits:
            required_by_split[split].add(item_id)

    resolved: dict[str, PerceptualDevelopmentItem] = {}
    locators: dict[str, tuple[str, int]] = {}
    datasets_by_split: dict[str, object] = {}
    rows_scanned: Counter[str] = Counter()

    for development_split, required in required_by_split.items():
        hf_split = split_names.get(development_split, development_split)
        dataset = load_dataset(dataset_name, "items", split=hf_split)
        missing_columns = {"item_id", "image"} - set(dataset.column_names)
        if missing_columns:
            raise KeyError(
                f"{dataset_name}/items/{hf_split} missing columns: "
                f"{sorted(missing_columns)}"
            )
        dataset = dataset.select_columns(["item_id", "image"])
        datasets_by_split[development_split] = dataset
        raw_item_ids = [str(value).strip() for value in dataset["item_id"]]
        rows_scanned[development_split] = len(raw_item_ids)
        for row_index, item_id in enumerate(raw_item_ids):
            if item_id not in required or item_id in resolved:
                continue
            row = dataset[row_index]
            fingerprint = _fingerprint_hf_value(row["image"])
            resolved[item_id] = PerceptualDevelopmentItem(
                item_id=item_id,
                splits=tuple(sorted(item_splits[item_id])),
                source_kit_ids=tuple(sorted(item_kits[item_id])),
                fingerprint=fingerprint,
                image_source=(
                    f"hf://{dataset_name}/items/{hf_split}/{item_id}"
                ),
            )
            locators[item_id] = (development_split, row_index)

    unresolved = sorted(set(item_splits) - set(resolved))
    store = HuggingFaceImageStore(
        datasets_by_split, locators, ssim_size=ssim_size
    )
    report = {
        "provider": "huggingface",
        "dataset": dataset_name,
        "split_mapping": dict(split_names),
        "rows_scanned": dict(rows_scanned),
        "resolved_item_count": len(resolved),
        "fingerprint_version": FINGERPRINT_VERSION,
    }
    return list(resolved.values()), store, report, unresolved


def _fit_preview(image, width: int, height: int):
    Image, ImageOps = _require_pillow()
    normalized = ImageOps.exif_transpose(image).convert("RGB")
    resampling = getattr(Image, "Resampling", Image)
    copy = normalized.copy()
    copy.thumbnail((width, height), resampling.LANCZOS)
    canvas = Image.new("RGB", (width, height), "white")
    x = (width - copy.width) // 2
    y = (height - copy.height) // 2
    canvas.paste(copy, (x, y))
    return canvas


def write_manual_preview(
    e3_path: Path,
    polyvore_image,
    pair_id: str,
    destination: Path,
) -> None:
    Image, _ = _require_pillow()
    from PIL import ImageDraw

    with Image.open(e3_path) as image:
        image.load()
        left = _fit_preview(image, 360, 360)
    right = _fit_preview(polyvore_image, 360, 360)
    canvas = Image.new("RGB", (760, 420), "white")
    canvas.paste(left, (10, 40))
    canvas.paste(right, (390, 40))
    draw = ImageDraw.Draw(canvas)
    draw.text((10, 10), f"{pair_id} | EVALUATION3", fill="black")
    draw.text((390, 10), "POLYVORE", fill="black")
    destination.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(destination, quality=92)


def _write_manual_html(
    rows: Sequence[Mapping[str, object]], destination: Path
) -> None:
    cards = []
    for row in rows:
        pair_id = html.escape(str(row["pair_id"]))
        preview = html.escape(str(row["preview_file"]))
        cards.append(
            f'<div class="card"><img src="{preview}" loading="lazy">'
            f"<div><b>{pair_id}</b></div></div>"
        )
    body = "\n".join(cards)
    text = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>E3 Manual Review</title>
<style>
body {{ font-family: sans-serif; margin: 20px; }}
.note {{ background:#f5f5f5; padding:12px; margin-bottom:18px; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(760px,1fr)); gap:18px; }}
.card {{ border:1px solid #bbb; padding:10px; }}
.card img {{ width:100%; height:auto; }}
</style></head><body>
<h1>EVALUATION3 pHash+SSIM — Manual Review</h1>
<div class="note">Fill <code>evaluation3_manual_review_BLIND.csv</code> using only <b>DUPLICATE</b> or <b>NON_DUPLICATE</b>. Do not use the KEY file while labeling.</div>
<div class="grid">{body}</div></body></html>"""
    destination.write_text(text, encoding="utf-8")


def _write_manual_xlsx(
    rows: Sequence[Mapping[str, object]], destination: Path
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required to write the manual-review workbook. "
            "Install requirements-evaluation.txt."
        ) from error

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "manual_review"
    headers = ["pair_id", "preview_file", "human_label"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get("pair_id", ""), row.get("preview_file", ""), ""])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 48
    sheet.column_dimensions["C"].width = 22
    validation = DataValidation(
        type="list",
        formula1='"DUPLICATE,NON_DUPLICATE"',
        allow_blank=True,
    )
    validation.error = "Chỉ chọn DUPLICATE hoặc NON_DUPLICATE."
    validation.errorTitle = "Invalid label"
    sheet.add_data_validation(validation)
    if rows:
        validation.add(f"C2:C{len(rows) + 1}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return [dict(row) for row in csv.DictReader(stream)]


def _read_manual_labels_file(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError(
                "openpyxl is required to read the manual-review workbook."
            ) from error
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        result: list[dict[str, str]] = []
        for values in rows[1:]:
            payload = {
                headers[index]: str(values[index] or "").strip()
                for index in range(min(len(headers), len(values)))
            }
            if payload.get("pair_id"):
                result.append(payload)
        return result
    return _read_csv(path)


def prepare_overlap_audit(
    *,
    evaluation3_root: Path | str,
    development_split_paths: Mapping[str, Path | str],
    output_dir: Path | str,
    polyvore_image_root: Path | str | None = None,
    polyvore_hf_dataset: str | None = None,
    hf_split_mapping: Mapping[str, str] | None = None,
    annotations_path: Path | str | None = None,
    annotation_sheet: str = "CMT",
    metadata_path: Path | str | None = None,
    metadata_sheet: str = "Num",
    evaluation3_groups: Iterable[str] | None = None,
    model_development_splits: Iterable[str] = DEFAULT_MODEL_DEVELOPMENT_SPLITS,
    use_exact_pixel: bool = False,
    phash_threshold: int = DEFAULT_PHASH_HAMMING_THRESHOLD,
    ssim_auto_threshold: float = DEFAULT_SSIM_AUTO_DUPLICATE_THRESHOLD,
    ssim_manual_lower_bound: float = DEFAULT_SSIM_MANUAL_LOWER_BOUND,
    ssim_size: int = DEFAULT_SSIM_SIZE,
    allow_incomplete_image_index: bool = False,
) -> tuple[dict[str, object], dict[str, Path]]:
    """Compute pHash+SSIM overlap evidence and emit a blind manual-review queue."""

    if bool(polyvore_image_root) == bool(polyvore_hf_dataset):
        raise ValueError("Choose exactly one Polyvore image provider")
    if not 0 <= phash_threshold <= 16:
        raise ValueError("phash_threshold must be between 0 and 16")
    if not 0 <= ssim_manual_lower_bound < ssim_auto_threshold <= 1:
        raise ValueError(
            "Require 0 <= manual SSIM lower bound < auto threshold <= 1"
        )

    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    e3_root = Path(evaluation3_root).expanduser().resolve()

    identity = load_development_identity(development_split_paths)
    if polyvore_image_root:
        development_items, store, provider_report, unresolved = (
            build_local_development_index(
                polyvore_image_root, identity, ssim_size=ssim_size
            )
        )
    else:
        development_items, store, provider_report, unresolved = (
            build_huggingface_development_index(
                str(polyvore_hf_dataset),
                hf_split_mapping or {},
                identity,
                ssim_size=ssim_size,
            )
        )
    if unresolved and not allow_incomplete_image_index:
        raise ValueError(
            f"Polyvore image index is incomplete: {len(unresolved)} unresolved "
            f"scorer item IDs; examples={unresolved[:10]}"
        )

    annotation_tables = []
    if annotations_path:
        annotation_tables.append(
            load_evaluation3_annotations(
                annotations_path, sheet_name=annotation_sheet
            )
        )
    if metadata_path:
        annotation_tables.append(
            load_evaluation3_annotations(metadata_path, sheet_name=metadata_sheet)
        )
    annotations = (
        merge_evaluation3_annotations(*annotation_tables)
        if annotation_tables
        else None
    )
    outfits = discover_evaluation3_outfits(
        e3_root,
        annotations=annotations,
        selected_groups=evaluation3_groups,
    )
    if not outfits:
        raise ValueError(
            "No EVALUATION3 outfit folders matched the requested selection"
        )

    exact_index: MutableMapping[
        str, list[PerceptualDevelopmentItem]
    ] = defaultdict(list)
    phash_index: MutableMapping[
        int, list[PerceptualDevelopmentItem]
    ] = defaultdict(list)
    tree = BKTree()
    for item in development_items:
        if use_exact_pixel:
            exact_index[item.fingerprint.pixel_sha256].append(item)
        if item.fingerprint.phash64 not in phash_index:
            tree.add(item.fingerprint.phash64)
        phash_index[item.fingerprint.phash64].append(item)

    audit_rows: list[dict[str, object]] = []
    auto_rows: list[dict[str, object]] = []
    exact_rows: list[dict[str, object]] = []
    manual_key_rows: list[dict[str, object]] = []
    manual_blind_rows: list[dict[str, object]] = []
    pair_counter = 0
    model_splits = set(model_development_splits)

    for outfit in outfits:
        outfit_id = str(outfit["e3_outfit_id"])
        auto_duplicate_splits: set[str] = set()
        manual_candidate_splits: set[str] = set()
        manual_pair_ids: list[str] = []
        image_results: list[dict[str, object]] = []

        for slot, relative_path in sorted(dict(outfit["images"]).items()):
            e3_path = e3_root / str(relative_path)
            Image, _ = _require_pillow()
            with Image.open(e3_path) as e3_image:
                e3_image.load()
                e3_copy = e3_image.copy()
            fingerprint = fingerprint_image(e3_copy)
            e3_ssim = image_to_ssim_array(e3_copy, size=ssim_size)
            matched_item_ids: set[str] = set()
            per_image_auto: list[dict[str, object]] = []
            per_image_manual: list[str] = []

            if use_exact_pixel:
                for reference in exact_index.get(
                    fingerprint.pixel_sha256, []
                ):
                    matched_item_ids.add(reference.item_id)
                    auto_duplicate_splits.update(reference.splits)
                    payload = {
                        "e3_outfit_id": outfit_id,
                        "e3_slot": slot,
                        "e3_image": str(relative_path),
                        "decision": DUPLICATE,
                        "method": "exact_pixel",
                        "polyvore_item_id": reference.item_id,
                        "polyvore_splits": list(reference.splits),
                        "polyvore_image_source": reference.image_source,
                        "phash_distance": 0,
                        "ssim": 1.0,
                    }
                    exact_rows.append(payload)
                    auto_rows.append(payload)
                    per_image_auto.append(payload)

            for phash_distance, phash_value in tree.query(
                fingerprint.phash64, phash_threshold
            ):
                for reference in phash_index[phash_value]:
                    if reference.item_id in matched_item_ids:
                        continue
                    poly_array = store.ssim_array(reference.item_id)
                    score = ssim_score(e3_ssim, poly_array)
                    decision, method = classify_pair(
                        phash_distance=phash_distance,
                        ssim=score,
                        exact_pixel_match=False,
                        use_exact_pixel=False,
                        phash_threshold=phash_threshold,
                        ssim_auto_threshold=ssim_auto_threshold,
                        ssim_manual_lower_bound=ssim_manual_lower_bound,
                    )
                    base = {
                        "e3_outfit_id": outfit_id,
                        "e3_slot": slot,
                        "e3_image": str(relative_path),
                        "decision": decision,
                        "method": method,
                        "polyvore_item_id": reference.item_id,
                        "polyvore_splits": list(reference.splits),
                        "polyvore_image_source": reference.image_source,
                        "phash_distance": int(phash_distance),
                        "ssim": float(score),
                    }
                    if decision == DUPLICATE:
                        auto_duplicate_splits.update(reference.splits)
                        auto_rows.append(base)
                        per_image_auto.append(base)
                    elif decision == MANUAL_REVIEW:
                        pair_counter += 1
                        pair_id = f"PAIR{pair_counter:06d}"
                        manual_candidate_splits.update(reference.splits)
                        manual_pair_ids.append(pair_id)
                        per_image_manual.append(pair_id)
                        preview_rel = (
                            f"manual_review_previews/{pair_id}.jpg"
                        )
                        preview_path = destination / preview_rel
                        write_manual_preview(
                            e3_path,
                            store.load_image(reference.item_id),
                            pair_id,
                            preview_path,
                        )
                        manual_key_rows.append(
                            {
                                "pair_id": pair_id,
                                **base,
                                "preview_file": preview_rel,
                            }
                        )
                        manual_blind_rows.append(
                            {
                                "pair_id": pair_id,
                                "preview_file": preview_rel,
                                "human_label": "",
                            }
                        )

            image_results.append(
                {
                    "slot": slot,
                    "evaluation3_image": str(relative_path),
                    "evaluation3_phash64": f"{fingerprint.phash64:016x}",
                    "evaluation3_pixel_sha256": (
                        fingerprint.pixel_sha256 if use_exact_pixel else None
                    ),
                    "auto_duplicate_count": len(per_image_auto),
                    "manual_review_pair_ids": per_image_manual,
                    "auto_duplicate_examples": per_image_auto[
                        :MAX_MATCH_EXAMPLES_PER_IMAGE
                    ],
                }
            )

        current_decision = (
            DUPLICATE
            if auto_duplicate_splits
            else MANUAL_REVIEW
            if manual_pair_ids
            else NON_DUPLICATE
        )
        audit_rows.append(
            {
                **dict(outfit),
                "overlap": {
                    "decision_pre_review": current_decision,
                    "auto_duplicate_splits": sorted(auto_duplicate_splits),
                    "manual_candidate_splits": sorted(manual_candidate_splits),
                    "manual_review_pair_ids": manual_pair_ids,
                    "image_results": image_results,
                },
            }
        )

    paths = {
        "summary": destination / "evaluation3_overlap_summary_pre_review.json",
        "audit_pre_review": (
            destination / "evaluation3_overlap_audit_pre_review.jsonl"
        ),
        "full": destination / "evaluation3_full.jsonl",
        "auto_duplicates": destination / "evaluation3_auto_duplicates.csv",
        "manual_blind": destination / "evaluation3_manual_review_BLIND.csv",
        "manual_xlsx": destination / "evaluation3_manual_review_BLIND.xlsx",
        "manual_key": destination / "evaluation3_manual_review_KEY.csv",
        "manual_html": destination / "evaluation3_manual_review_BLIND.html",
    }
    if use_exact_pixel:
        paths["exact_pixel_duplicates"] = (
            destination / "evaluation3_exact_pixel_duplicates.csv"
        )

    write_jsonl(audit_rows, paths["audit_pre_review"])
    write_jsonl(audit_rows, paths["full"])

    evidence_headers = (
        "e3_outfit_id",
        "e3_slot",
        "e3_image",
        "decision",
        "method",
        "polyvore_item_id",
        "polyvore_splits",
        "polyvore_image_source",
        "phash_distance",
        "ssim",
    )
    with paths["auto_duplicates"].open(
        "w", encoding="utf-8", newline=""
    ) as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=evidence_headers,
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in auto_rows:
            serialized = dict(row)
            serialized["polyvore_splits"] = "|".join(
                serialized.get("polyvore_splits", [])
            )
            writer.writerow(serialized)
    if use_exact_pixel:
        with paths["exact_pixel_duplicates"].open(
            "w", encoding="utf-8", newline=""
        ) as stream:
            writer = csv.DictWriter(
                stream,
                fieldnames=evidence_headers,
                extrasaction="ignore",
            )
            writer.writeheader()
            for row in exact_rows:
                serialized = dict(row)
                serialized["polyvore_splits"] = "|".join(
                    serialized.get("polyvore_splits", [])
                )
                writer.writerow(serialized)

    blind_headers = ("pair_id", "preview_file", "human_label")
    with paths["manual_blind"].open(
        "w", encoding="utf-8", newline=""
    ) as stream:
        writer = csv.DictWriter(stream, fieldnames=blind_headers)
        writer.writeheader()
        writer.writerows(manual_blind_rows)

    key_headers = (
        "pair_id",
        "preview_file",
        "e3_outfit_id",
        "e3_slot",
        "e3_image",
        "polyvore_item_id",
        "polyvore_splits",
        "polyvore_image_source",
        "phash_distance",
        "ssim",
        "decision",
        "method",
    )
    with paths["manual_key"].open(
        "w", encoding="utf-8", newline=""
    ) as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=key_headers,
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in manual_key_rows:
            serialized = dict(row)
            serialized["polyvore_splits"] = "|".join(
                serialized.get("polyvore_splits", [])
            )
            writer.writerow(serialized)
    _write_manual_html(manual_blind_rows, paths["manual_html"])
    _write_manual_xlsx(manual_blind_rows, paths["manual_xlsx"])

    incomplete_e3 = sum(bool(row.get("missing_slots")) for row in audit_rows)
    summary = {
        "status": (
            "MANUAL_REVIEW_REQUIRED"
            if manual_blind_rows
            else "PRE_REVIEW_COMPLETE"
        ),
        "official_clean_manifests_ready": False,
        "configuration": {
            "protocol": "phash-ssim-v1",
            "use_exact_pixel_shortcut": bool(use_exact_pixel),
            "id_collision_policy": "not_used",
            "phash_hamming_threshold": phash_threshold,
            "ssim_auto_duplicate_threshold": ssim_auto_threshold,
            "ssim_manual_lower_bound": ssim_manual_lower_bound,
            "ssim_preprocess": (
                f"EXIF transpose -> grayscale -> {ssim_size}x{ssim_size} "
                "LANCZOS"
            ),
            "phash": (
                f"DCT pHash {PHASH_HASH_SIZE * PHASH_HASH_SIZE}-bit, "
                f"hash_size={PHASH_HASH_SIZE}, "
                f"highfreq_factor={PHASH_HIGHFREQ_FACTOR}"
            ),
            "model_development_splits": sorted(model_splits),
        },
        "development_data": {
            "sample_counts": identity["sample_counts"],
            "unique_item_count": len(identity["item_splits"]),
            "fingerprinted_item_count": len(development_items),
            "unresolved_item_count": len(unresolved),
            "unresolved_item_id_examples": unresolved[:50],
            "image_provider": provider_report,
        },
        "evaluation3": {
            "selected_outfit_count": len(audit_rows),
            "outfits_missing_required_images": incomplete_e3,
        },
        "decisions_pre_review": {
            "auto_duplicate_pairs": len(auto_rows),
            "exact_pixel_pairs": len(exact_rows),
            "manual_review_pairs": len(manual_blind_rows),
            "outfits_auto_duplicate": sum(
                row["overlap"]["decision_pre_review"] == DUPLICATE
                for row in audit_rows
            ),
            "outfits_manual_review": sum(
                row["overlap"]["decision_pre_review"] == MANUAL_REVIEW
                for row in audit_rows
            ),
            "outfits_non_duplicate": sum(
                row["overlap"]["decision_pre_review"] == NON_DUPLICATE
                for row in audit_rows
            ),
        },
        "next_step": (
            "Fill evaluation3_manual_review_BLIND.xlsx with "
            "DUPLICATE/NON_DUPLICATE, then run finalize_overlap_audit."
            if manual_blind_rows
            else "Run finalize_overlap_audit; no manual labels are required."
        ),
    }
    write_json(summary, paths["summary"])
    return summary, paths


def finalize_overlap_audit(
    *,
    output_dir: Path | str,
    manual_labels_path: Path | str | None = None,
    model_development_splits: Iterable[str] = DEFAULT_MODEL_DEVELOPMENT_SPLITS,
) -> tuple[dict[str, object], dict[str, Path]]:
    """Apply binary manual labels and write final clean manifests."""

    destination = Path(output_dir).expanduser().resolve()
    pre_audit_path = destination / "evaluation3_overlap_audit_pre_review.jsonl"
    default_csv_path = destination / "evaluation3_manual_review_BLIND.csv"
    default_xlsx_path = destination / "evaluation3_manual_review_BLIND.xlsx"
    blind_path = (
        Path(manual_labels_path).expanduser().resolve()
        if manual_labels_path is not None
        else default_xlsx_path
        if default_xlsx_path.is_file()
        else default_csv_path
    )
    key_path = destination / "evaluation3_manual_review_KEY.csv"
    pre_summary_path = destination / "evaluation3_overlap_summary_pre_review.json"
    if (
        not pre_audit_path.is_file()
        or not blind_path.is_file()
        or not key_path.is_file()
    ):
        raise FileNotFoundError(
            "Missing pre-review outputs; run prepare_overlap_audit first"
        )

    audit_rows = [
        json.loads(line)
        for line in pre_audit_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    blind_rows = _read_manual_labels_file(blind_path)
    key_rows = _read_csv(key_path)
    labels: dict[str, str] = {}
    invalid: list[tuple[str, str]] = []
    unresolved_pairs: list[str] = []
    for row in blind_rows:
        pair_id = str(row.get("pair_id", "")).strip()
        label = str(row.get("human_label", "")).strip().upper()
        if not label:
            unresolved_pairs.append(pair_id)
            continue
        if label not in VALID_FINAL_MANUAL_LABELS:
            invalid.append((pair_id, label))
        else:
            labels[pair_id] = label
    if invalid:
        raise ValueError(
            "Manual review accepts only DUPLICATE or NON_DUPLICATE. "
            f"Invalid examples: {invalid[:10]}"
        )

    key_by_pair = {str(row["pair_id"]): row for row in key_rows}
    model_splits = set(model_development_splits)
    final_rows: list[dict[str, object]] = []
    for row in audit_rows:
        overlap = dict(row["overlap"])
        confirmed_splits = set(overlap.get("auto_duplicate_splits", []))
        row_unresolved: list[str] = []
        manual_duplicate_pair_ids: list[str] = []
        manual_non_duplicate_pair_ids: list[str] = []
        for pair_id in overlap.get("manual_review_pair_ids", []):
            label = labels.get(pair_id)
            if label is None:
                row_unresolved.append(pair_id)
                continue
            if label == DUPLICATE:
                manual_duplicate_pair_ids.append(pair_id)
                key = key_by_pair[pair_id]
                confirmed_splits.update(
                    value
                    for value in str(key.get("polyvore_splits", "")).split("|")
                    if value
                )
            else:
                manual_non_duplicate_pair_ids.append(pair_id)

        final_decision = (
            DUPLICATE
            if confirmed_splits
            else MANUAL_REVIEW
            if row_unresolved
            else NON_DUPLICATE
        )
        final_overlap = {
            **overlap,
            "decision_final": final_decision,
            "confirmed_duplicate_splits": sorted(confirmed_splits),
            "manual_duplicate_pair_ids": manual_duplicate_pair_ids,
            "manual_non_duplicate_pair_ids": manual_non_duplicate_pair_ids,
            "manual_unresolved_pair_ids": row_unresolved,
        }
        row_has_missing_images = bool(row.get("missing_slots"))
        final_rows.append(
            {
                **row,
                "overlap": final_overlap,
                "eligible_for_full": True,
                "eligible_for_model_clean": (
                    (
                        final_decision == NON_DUPLICATE
                        or (
                            final_decision == DUPLICATE
                            and not (confirmed_splits & model_splits)
                        )
                    )
                    if not row_unresolved and not row_has_missing_images
                    else False
                ),
                "eligible_for_strict_clean": (
                    final_decision == NON_DUPLICATE
                    if not row_unresolved and not row_has_missing_images
                    else False
                ),
            }
        )

    pre_summary = (
        json.loads(pre_summary_path.read_text(encoding="utf-8"))
        if pre_summary_path.is_file()
        else {}
    )
    incomplete_e3 = int(
        pre_summary.get("evaluation3", {}).get(
            "outfits_missing_required_images", 0
        )
    )
    unresolved_dev = int(
        pre_summary.get("development_data", {}).get("unresolved_item_count", 0)
    )
    ready = not unresolved_pairs and incomplete_e3 == 0 and unresolved_dev == 0
    summary = {
        **pre_summary,
        "status": "PASS" if ready else "INCOMPLETE",
        "official_clean_manifests_ready": ready,
        "manual_review": {
            "total_pairs": len(blind_rows),
            "duplicate": sum(
                label == DUPLICATE for label in labels.values()
            ),
            "non_duplicate": sum(
                label == NON_DUPLICATE for label in labels.values()
            ),
            "remaining": len(unresolved_pairs),
            "remaining_examples": unresolved_pairs[:50],
        },
        "manifests": {
            "full": len(final_rows),
            "model_clean": sum(
                bool(row["eligible_for_model_clean"]) for row in final_rows
            ),
            "strict_clean": sum(
                bool(row["eligible_for_strict_clean"]) for row in final_rows
            ),
        },
    }

    paths = {
        "summary": destination / "evaluation3_overlap_summary.json",
        "audit": destination / "evaluation3_overlap_audit.jsonl",
        "full": destination / "evaluation3_full.jsonl",
        "manual_labels": blind_path,
        "manual_key": key_path,
    }
    write_json(summary, paths["summary"])
    write_jsonl(final_rows, paths["audit"])
    write_jsonl(final_rows, paths["full"])
    if ready:
        paths["model_clean"] = destination / "evaluation3_model_clean.jsonl"
        paths["strict_clean"] = destination / "evaluation3_strict_clean.jsonl"
        write_jsonl(
            (
                row
                for row in final_rows
                if row["eligible_for_model_clean"]
            ),
            paths["model_clean"],
        )
        write_jsonl(
            (
                row
                for row in final_rows
                if row["eligible_for_strict_clean"]
            ),
            paths["strict_clean"],
        )
    return summary, paths

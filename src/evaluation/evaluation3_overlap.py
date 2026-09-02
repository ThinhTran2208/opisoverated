# -*- coding: utf-8 -*-
"""Audit image overlap between EVALUATION3 and the Polyvore scorer data.

The scorer-ready JSONL contains the exact item IDs presented to the model but
not the item images. This module joins those IDs to either a local Polyvore
image directory or the Hugging Face ``codewaly/polyvore1000`` dataset, builds
image fingerprints, and checks every EVALUATION3 U/B/S/G image against them.

Three manifests are written:

``full``
    Every selected EVALUATION3 outfit. This is diagnostic only when overlap is
    present.
``model_clean``
    No overlap with the configured model-development splits (normally train
    and valid).
``strict_clean``
    No overlap with any supplied Polyvore split, including test.

An ID collision is reported separately from an image match because expanded
EVALUATION3 combines more than one source dataset. The clean manifests use a
conservative rule and exclude ID candidates, exact pixel duplicates, and
near-duplicate image candidates.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Mapping, MutableMapping, Sequence


IMAGE_SUFFIXES = frozenset({".jpg", ".jpeg", ".png", ".webp", ".bmp"})
EVALUATION3_SLOTS = {
    "U": "top",
    "B": "bottom",
    "S": "shoes",
    "G": "bag",
}
DEFAULT_MODEL_DEVELOPMENT_SPLITS = ("train", "valid")
DEFAULT_NEAR_HAMMING_THRESHOLD = 4
MAX_MATCH_EXAMPLES_PER_IMAGE = 20


@dataclass(frozen=True)
class ImageFingerprint:
    """Encoding-independent exact pixels plus a resize-tolerant dHash."""

    pixel_sha256: str
    dhash64: int
    width: int
    height: int


@dataclass(frozen=True)
class DevelopmentItem:
    item_id: str
    splits: tuple[str, ...]
    source_kit_ids: tuple[str, ...]
    fingerprint: ImageFingerprint
    image_source: str


class BKTree:
    """Small Hamming-distance index for 64-bit perceptual hashes."""

    def __init__(self) -> None:
        self._root: tuple[int, dict] | None = None

    def add(self, value: int) -> None:
        if self._root is None:
            self._root = (value, {})
            return

        node = self._root
        while True:
            node_value, children = node
            distance = hamming_distance(value, node_value)
            child = children.get(distance)
            if child is None:
                children[distance] = (value, {})
                return
            node = child

    def query(self, value: int, max_distance: int) -> list[tuple[int, int]]:
        if self._root is None:
            return []

        matches: list[tuple[int, int]] = []
        pending = [self._root]
        while pending:
            node_value, children = pending.pop()
            distance = hamming_distance(value, node_value)
            if distance <= max_distance:
                matches.append((distance, node_value))
            lower = distance - max_distance
            upper = distance + max_distance
            pending.extend(
                child
                for edge, child in children.items()
                if lower <= edge <= upper
            )
        return sorted(matches)


def hamming_distance(left: int, right: int) -> int:
    return (left ^ right).bit_count()


def _require_pillow():
    try:
        from PIL import Image, ImageOps
    except ImportError as error:
        raise RuntimeError(
            "Pillow is required for the overlap audit. Install "
            "requirements-evaluation.txt."
        ) from error
    return Image, ImageOps


def fingerprint_image(image) -> ImageFingerprint:
    """Fingerprint a Pillow image after EXIF orientation and RGB decoding."""

    Image, ImageOps = _require_pillow()
    normalized = ImageOps.exif_transpose(image).convert("RGB")
    width, height = normalized.size

    exact = hashlib.sha256()
    exact.update(f"RGB:{width}x{height}:".encode("ascii"))
    exact.update(normalized.tobytes())

    # dHash compares adjacent luminance samples. It is intentionally used as
    # a candidate detector, not proof of identity.
    resampling = getattr(Image, "Resampling", Image)
    small = normalized.convert("L").resize((9, 8), resampling.LANCZOS)
    get_pixels = getattr(small, "get_flattened_data", small.getdata)
    pixels = list(get_pixels())
    dhash = 0
    for row in range(8):
        offset = row * 9
        for column in range(8):
            dhash = (dhash << 1) | int(
                pixels[offset + column] > pixels[offset + column + 1]
            )

    return ImageFingerprint(
        pixel_sha256=exact.hexdigest(),
        dhash64=dhash,
        width=width,
        height=height,
    )


def fingerprint_path(path: Path | str) -> ImageFingerprint:
    Image, _ = _require_pillow()
    source = Path(path)
    try:
        with Image.open(source) as image:
            image.load()
            return fingerprint_image(image)
    except Exception as error:
        raise ValueError(f"Cannot decode image: {source}: {error}") from error


def read_jsonl(path: Path | str) -> list[dict]:
    source = Path(path)
    rows: list[dict] = []
    with source.open("r", encoding="utf-8") as stream:
        for line_number, raw_line in enumerate(stream, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as error:
                raise ValueError(
                    f"Invalid JSON at {source}:{line_number}: {error}"
                ) from error
            if not isinstance(row, dict):
                raise ValueError(f"Expected JSON object at {source}:{line_number}")
            rows.append(row)
    return rows


def write_json(payload: Mapping[str, object], path: Path | str) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2)
        stream.write("\n")


def write_jsonl(rows: Iterable[Mapping[str, object]], path: Path | str) -> int:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with destination.open("w", encoding="utf-8") as stream:
        for row in rows:
            stream.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
            stream.write("\n")
            count += 1
    return count


def parse_named_paths(values: Sequence[str], option_name: str) -> dict[str, Path]:
    """Parse repeated ``NAME=PATH`` CLI values."""

    result: dict[str, Path] = {}
    for value in values:
        name, separator, raw_path = value.partition("=")
        name = name.strip()
        raw_path = raw_path.strip()
        if not separator or not name or not raw_path:
            raise ValueError(f"{option_name} must use NAME=PATH, got {value!r}")
        if name in result:
            raise ValueError(f"Duplicate {option_name} name: {name}")
        path = Path(raw_path).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"{option_name} file not found: {path}")
        result[name] = path
    if not result:
        raise ValueError(f"At least one {option_name} is required")
    return result


def load_development_identity(
    split_paths: Mapping[str, Path | str],
) -> dict[str, object]:
    """Collect every item actually presented to the scorer in each split."""

    item_splits: MutableMapping[str, set[str]] = defaultdict(set)
    item_kits: MutableMapping[str, set[str]] = defaultdict(set)
    kit_splits: MutableMapping[str, set[str]] = defaultdict(set)
    sample_counts: Counter[str] = Counter()

    for split, path in split_paths.items():
        for row_number, row in enumerate(read_jsonl(path), start=1):
            sample_id = str(row.get("sample_id", "")).strip()
            kit_id = str(row.get("source_kit_id", "")).strip()
            raw_items = row.get("items")
            if not sample_id or not kit_id or not isinstance(raw_items, list):
                raise ValueError(
                    f"Invalid scorer row {path}:{row_number}; expected sample_id, "
                    "source_kit_id and items[]"
                )
            if not raw_items:
                raise ValueError(f"Empty items at {path}:{row_number}")

            kit_splits[kit_id].add(split)
            sample_counts[split] += 1
            for raw_item in raw_items:
                item_id = str(raw_item).strip()
                if not item_id:
                    raise ValueError(f"Empty item ID at {path}:{row_number}")
                item_splits[item_id].add(split)
                item_kits[item_id].add(kit_id)

    return {
        "item_splits": {key: set(value) for key, value in item_splits.items()},
        "item_kits": {key: set(value) for key, value in item_kits.items()},
        "kit_splits": {key: set(value) for key, value in kit_splits.items()},
        "sample_counts": dict(sample_counts),
    }


def _candidate_item_ids(path: Path, root: Path) -> tuple[str, ...]:
    relative = path.relative_to(root)
    values = {path.stem}
    if len(relative.parts) >= 2:
        values.add(f"{relative.parts[-2]}_{path.stem}")
    return tuple(values)


def fingerprint_local_development_images(
    image_root: Path | str,
    required_item_ids: Iterable[str],
) -> tuple[dict[str, tuple[ImageFingerprint, str]], dict[str, object]]:
    """Resolve required item IDs from file stems or ``kit/slot`` paths."""

    root = Path(image_root).expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"Polyvore image root not found: {root}")

    required = set(required_item_ids)
    resolved: dict[str, tuple[ImageFingerprint, str]] = {}
    duplicate_candidates: MutableMapping[str, list[str]] = defaultdict(list)
    scanned = 0

    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_SUFFIXES:
            continue
        scanned += 1
        matched = required.intersection(_candidate_item_ids(path, root))
        for item_id in matched:
            if item_id in resolved:
                duplicate_candidates[item_id].append(str(path))
                continue
            fingerprint = fingerprint_path(path)
            resolved[item_id] = (fingerprint, str(path))

    return resolved, {
        "provider": "local_directory",
        "root": str(root),
        "files_scanned": scanned,
        "resolved_item_count": len(resolved),
        "duplicate_path_item_count": len(duplicate_candidates),
        "duplicate_path_examples": {
            key: value[:5]
            for key, value in list(sorted(duplicate_candidates.items()))[:20]
        },
    }


def _fingerprint_hf_value(value) -> ImageFingerprint:
    Image, _ = _require_pillow()
    if isinstance(value, Image.Image):
        return fingerprint_image(value)
    if isinstance(value, Mapping):
        if value.get("bytes") is not None:
            from io import BytesIO

            with Image.open(BytesIO(value["bytes"])) as image:
                image.load()
                return fingerprint_image(image)
        if value.get("path"):
            return fingerprint_path(value["path"])
    if isinstance(value, (str, Path)):
        return fingerprint_path(value)
    raise ValueError(f"Unsupported Hugging Face image value: {type(value).__name__}")


def fingerprint_huggingface_development_images(
    dataset_name: str,
    split_names: Mapping[str, str],
    item_splits: Mapping[str, set[str]],
) -> tuple[dict[str, tuple[ImageFingerprint, str]], dict[str, object]]:
    """Read only needed item images from a Hugging Face dataset."""

    try:
        from datasets import load_dataset
    except ImportError as error:
        raise RuntimeError(
            "datasets is required for --polyvore-hf-dataset. Install "
            "requirements-evaluation.txt."
        ) from error

    required_by_split: MutableMapping[str, set[str]] = defaultdict(set)
    for item_id, development_splits in item_splits.items():
        for development_split in development_splits:
            required_by_split[development_split].add(item_id)

    resolved: dict[str, tuple[ImageFingerprint, str]] = {}
    rows_scanned: Counter[str] = Counter()
    for development_split, required in required_by_split.items():
        hf_split = split_names.get(development_split, development_split)
        dataset = load_dataset(dataset_name, "items", split=hf_split)
        missing_columns = {"item_id", "image"} - set(dataset.column_names)
        if missing_columns:
            raise KeyError(
                f"{dataset_name}/items/{hf_split} missing columns: "
                f"{sorted(missing_columns)}"
            )
        dataset = dataset.select_columns(["item_id", "image"])
        raw_item_ids = [str(value).strip() for value in dataset["item_id"]]
        selected_indices = [
            index
            for index, item_id in enumerate(raw_item_ids)
            if item_id in required and item_id not in resolved
        ]
        rows_scanned[development_split] = len(raw_item_ids)
        for row in dataset.select(selected_indices):
            item_id = str(row.get("item_id", "")).strip()
            resolved[item_id] = (
                _fingerprint_hf_value(row["image"]),
                f"hf://{dataset_name}/items/{hf_split}/{item_id}",
            )

    return resolved, {
        "provider": "huggingface",
        "dataset": dataset_name,
        "split_mapping": dict(split_names),
        "rows_scanned": dict(rows_scanned),
        "selected_image_rows": len(resolved),
        "resolved_item_count": len(resolved),
    }


def build_development_items(
    identity: Mapping[str, object],
    image_fingerprints: Mapping[str, tuple[ImageFingerprint, str]],
) -> list[DevelopmentItem]:
    item_splits = identity["item_splits"]
    item_kits = identity["item_kits"]
    return [
        DevelopmentItem(
            item_id=item_id,
            splits=tuple(sorted(item_splits[item_id])),
            source_kit_ids=tuple(sorted(item_kits[item_id])),
            fingerprint=fingerprint,
            image_source=image_source,
        )
        for item_id, (fingerprint, image_source) in sorted(image_fingerprints.items())
    ]


def normalize_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def load_evaluation3_annotations(
    path: Path | str,
    *,
    sheet_name: str = "CMT",
) -> dict[str, dict[str, object]]:
    """Load ITEM/Cmt/Reason/Group from CSV or the supplied workbook."""

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"EVALUATION3 annotation file not found: {source}")

    if source.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.reader(stream, delimiter=delimiter))
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError(
                "openpyxl is required to read EVALUATION3 annotations. Install "
                "requirements-evaluation.txt."
            ) from error
        # A Drive download may lose the .xlsx suffix. Passing a binary stream
        # lets openpyxl inspect the ZIP content instead of rejecting the name.
        with source.open("rb") as workbook_stream:
            workbook = load_workbook(workbook_stream, read_only=True, data_only=True)
            selected = next(
                (
                    name
                    for name in workbook.sheetnames
                    if name.lower() == sheet_name.lower()
                ),
                None,
            )
            if selected is None:
                raise ValueError(
                    f"Sheet {sheet_name!r} not found in {source}; "
                    f"available={workbook.sheetnames}"
                )
            rows = [
                list(row)
                for row in workbook[selected].iter_rows(values_only=True)
            ]
            workbook.close()

    if not rows:
        raise ValueError(f"Annotation table is empty: {source}")
    headers = [_normalized_header(value) for value in rows[0]]
    aliases = {
        "item": {"item", "item#", "outfit_id", "outfitid"},
        "cmt": {"cmt", "judgment", "judgement", "label"},
        "reason": {"reason", "decisive_reason"},
        "group": {"group", "split"},
    }
    columns: dict[str, int] = {}
    for target, candidates in aliases.items():
        for index, header in enumerate(headers):
            if header in candidates:
                columns[target] = index
                break
    if "item" not in columns:
        raise ValueError(f"Annotation table has no ITEM/outfit_id column: {headers}")

    annotations: dict[str, dict[str, object]] = {}
    for row_number, row in enumerate(rows[1:], start=2):
        outfit_id = normalize_id(row[columns["item"]] if columns["item"] < len(row) else None)
        if not outfit_id:
            continue
        if outfit_id in annotations:
            raise ValueError(f"Duplicate annotation ITEM {outfit_id} at row {row_number}")
        annotations[outfit_id] = {
            "cmt_raw": normalize_id(row[columns["cmt"]]) if "cmt" in columns else None,
            "reason_raw": normalize_id(row[columns["reason"]]) if "reason" in columns else None,
            "group": normalize_id(row[columns["group"]]) if "group" in columns else None,
        }
    return annotations


def merge_evaluation3_annotations(
    *tables: Mapping[str, Mapping[str, object]],
) -> dict[str, dict[str, object]]:
    """Join Cmt and attribute tables by ITEM while rejecting conflicts."""

    merged: dict[str, dict[str, object]] = {}
    for table in tables:
        for outfit_id, row in table.items():
            target = merged.setdefault(
                outfit_id,
                {"cmt_raw": None, "reason_raw": None, "group": None},
            )
            for field in ("cmt_raw", "reason_raw", "group"):
                incoming = row.get(field)
                if incoming in (None, ""):
                    continue
                existing = target.get(field)
                if existing not in (None, "") and existing != incoming:
                    raise ValueError(
                        f"Conflicting EVALUATION3 {field} for ITEM {outfit_id}: "
                        f"{existing!r} vs {incoming!r}"
                    )
                target[field] = incoming
    return merged


def discover_evaluation3_outfits(
    image_root: Path | str,
    *,
    annotations: Mapping[str, Mapping[str, object]] | None = None,
    selected_groups: Iterable[str] | None = None,
) -> list[dict[str, object]]:
    """Find U/B/S/G files under each immediate EVALUATION3 outfit folder."""

    root = Path(image_root).expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"EVALUATION3 image root not found: {root}")
    allowed_groups = {value.strip() for value in selected_groups or [] if value.strip()}
    if allowed_groups and annotations is None:
        raise ValueError("--evaluation3-group requires --evaluation3-annotations")

    outfits: list[dict[str, object]] = []
    for directory in sorted((path for path in root.iterdir() if path.is_dir()), key=lambda p: p.name):
        outfit_id = normalize_id(directory.name)
        annotation = dict((annotations or {}).get(outfit_id, {}))
        if allowed_groups and annotation.get("group") not in allowed_groups:
            continue

        slot_paths: dict[str, str] = {}
        duplicate_slots: MutableMapping[str, list[str]] = defaultdict(list)
        for path in sorted(directory.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in IMAGE_SUFFIXES:
                continue
            semantic_slot = EVALUATION3_SLOTS.get(path.stem.upper())
            if semantic_slot is None:
                continue
            relative = str(path.relative_to(root))
            if semantic_slot in slot_paths:
                duplicate_slots[semantic_slot].append(relative)
            else:
                slot_paths[semantic_slot] = relative

        outfits.append(
            {
                "e3_outfit_id": outfit_id,
                "group": annotation.get("group"),
                "cmt_raw": annotation.get("cmt_raw"),
                "reason_raw": annotation.get("reason_raw"),
                "images": slot_paths,
                "missing_slots": sorted(set(EVALUATION3_SLOTS.values()) - set(slot_paths)),
                "duplicate_slots": dict(duplicate_slots),
            }
        )
    return outfits


def _reference_payload(reference: DevelopmentItem, distance: int | None = None) -> dict:
    payload = {
        "polyvore_item_id": reference.item_id,
        "polyvore_splits": list(reference.splits),
        "polyvore_source_kit_ids": list(reference.source_kit_ids),
        "polyvore_image_source": reference.image_source,
    }
    if distance is not None:
        payload["dhash_hamming_distance"] = distance
    return payload


def audit_overlap(
    evaluation3_root: Path | str,
    outfits: Sequence[Mapping[str, object]],
    identity: Mapping[str, object],
    development_items: Sequence[DevelopmentItem],
    *,
    model_development_splits: Iterable[str] = DEFAULT_MODEL_DEVELOPMENT_SPLITS,
    near_hamming_threshold: int = DEFAULT_NEAR_HAMMING_THRESHOLD,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Return outfit audit rows and a flat list of image-match evidence."""

    if not 0 <= near_hamming_threshold <= 16:
        raise ValueError("near_hamming_threshold must be between 0 and 16")

    e3_root = Path(evaluation3_root).expanduser().resolve()
    exact_index: MutableMapping[str, list[DevelopmentItem]] = defaultdict(list)
    dhash_index: MutableMapping[int, list[DevelopmentItem]] = defaultdict(list)
    tree = BKTree()
    for item in development_items:
        exact_index[item.fingerprint.pixel_sha256].append(item)
        if item.fingerprint.dhash64 not in dhash_index:
            tree.add(item.fingerprint.dhash64)
        dhash_index[item.fingerprint.dhash64].append(item)

    model_splits = set(model_development_splits)
    kit_splits = identity["kit_splits"]
    audit_rows: list[dict[str, object]] = []
    evidence_rows: list[dict[str, object]] = []

    for outfit in outfits:
        outfit_id = str(outfit["e3_outfit_id"])
        id_candidate_splits = set(kit_splits.get(outfit_id, set()))
        exact_splits: set[str] = set()
        near_splits: set[str] = set()
        image_matches: list[dict[str, object]] = []

        for slot, relative_path in sorted(dict(outfit["images"]).items()):
            e3_path = e3_root / str(relative_path)
            fingerprint = fingerprint_path(e3_path)
            exact_references = exact_index.get(fingerprint.pixel_sha256, [])
            exact_ids = {reference.item_id for reference in exact_references}
            exact_payload = [
                _reference_payload(reference)
                for reference in exact_references[:MAX_MATCH_EXAMPLES_PER_IMAGE]
            ]
            for reference in exact_references:
                exact_splits.update(reference.splits)

            near_candidates: list[tuple[int, DevelopmentItem]] = []
            if near_hamming_threshold > 0:
                for distance, dhash in tree.query(
                    fingerprint.dhash64, near_hamming_threshold
                ):
                    for reference in dhash_index[dhash]:
                        if reference.item_id not in exact_ids:
                            near_candidates.append((distance, reference))
                near_candidates.sort(key=lambda value: (value[0], value[1].item_id))
                for _, reference in near_candidates:
                    near_splits.update(reference.splits)

            near_payload = [
                _reference_payload(reference, distance)
                for distance, reference in near_candidates[:MAX_MATCH_EXAMPLES_PER_IMAGE]
            ]
            match = {
                "slot": slot,
                "evaluation3_image": str(relative_path),
                "evaluation3_pixel_sha256": fingerprint.pixel_sha256,
                "evaluation3_dhash64": f"{fingerprint.dhash64:016x}",
                "exact_matches": exact_payload,
                "exact_match_count": len(exact_references),
                "near_matches": near_payload,
                "near_match_count": len(near_candidates),
            }
            image_matches.append(match)

            for match_type, payloads in (
                ("exact_pixel", exact_payload),
                ("near_dhash", near_payload),
            ):
                for payload in payloads:
                    evidence_rows.append(
                        {
                            "e3_outfit_id": outfit_id,
                            "e3_slot": slot,
                            "e3_image": str(relative_path),
                            "match_type": match_type,
                            **payload,
                        }
                    )

        all_overlap_splits = id_candidate_splits | exact_splits | near_splits
        overlap_model_splits = all_overlap_splits & model_splits
        audit_rows.append(
            {
                **dict(outfit),
                "overlap": {
                    "id_candidate_splits": sorted(id_candidate_splits),
                    "exact_pixel_splits": sorted(exact_splits),
                    "near_dhash_candidate_splits": sorted(near_splits),
                    "all_overlap_splits": sorted(all_overlap_splits),
                    "model_development_overlap_splits": sorted(overlap_model_splits),
                    "image_matches": image_matches,
                },
                "eligible_for_full": True,
                "eligible_for_model_clean": not overlap_model_splits,
                "eligible_for_strict_clean": not all_overlap_splits,
            }
        )

    return audit_rows, evidence_rows


def build_summary(
    audit_rows: Sequence[Mapping[str, object]],
    identity: Mapping[str, object],
    image_provider_report: Mapping[str, object],
    *,
    unresolved_development_item_ids: Sequence[str],
    model_development_splits: Iterable[str],
    near_hamming_threshold: int,
) -> dict[str, object]:
    method_counts = Counter()
    split_counts = Counter()
    label_counts = Counter()
    group_counts = Counter()
    incomplete_e3 = 0
    for row in audit_rows:
        label_counts[str(row.get("cmt_raw"))] += 1
        group_counts[str(row.get("group"))] += 1
        if row.get("missing_slots"):
            incomplete_e3 += 1
        overlap = row["overlap"]
        for method in (
            "id_candidate_splits",
            "exact_pixel_splits",
            "near_dhash_candidate_splits",
        ):
            if overlap[method]:
                method_counts[method] += 1
        for split in overlap["all_overlap_splits"]:
            split_counts[split] += 1

    total_development_items = len(identity["item_splits"])
    image_index_complete = not unresolved_development_item_ids
    e3_images_complete = incomplete_e3 == 0
    official_clean_manifests_ready = image_index_complete and e3_images_complete
    return {
        "status": "PASS" if official_clean_manifests_ready else "INCOMPLETE",
        "official_clean_manifests_ready": official_clean_manifests_ready,
        "configuration": {
            "model_development_splits": sorted(set(model_development_splits)),
            "near_dhash_hamming_threshold": near_hamming_threshold,
            "clean_policy": (
                "exclude ID candidates, exact pixel duplicates, and near-dHash "
                "candidates in the relevant Polyvore splits"
            ),
        },
        "development_data": {
            "sample_counts": identity["sample_counts"],
            "source_kit_count": len(identity["kit_splits"]),
            "unique_item_count": total_development_items,
            "fingerprinted_item_count": total_development_items
            - len(unresolved_development_item_ids),
            "unresolved_item_count": len(unresolved_development_item_ids),
            "unresolved_item_id_examples": list(unresolved_development_item_ids[:50]),
            "image_provider": dict(image_provider_report),
        },
        "evaluation3": {
            "selected_outfit_count": len(audit_rows),
            "raw_cmt_counts": dict(sorted(label_counts.items())),
            "group_counts": dict(sorted(group_counts.items())),
            "outfits_missing_required_images": incomplete_e3,
        },
        "overlap": {
            "outfit_counts_by_method": dict(method_counts),
            "outfit_counts_by_polyvore_split": dict(sorted(split_counts.items())),
        },
        "manifests": {
            "full": sum(bool(row["eligible_for_full"]) for row in audit_rows),
            "model_clean": sum(
                bool(row["eligible_for_model_clean"]) for row in audit_rows
            ),
            "strict_clean": sum(
                bool(row["eligible_for_strict_clean"]) for row in audit_rows
            ),
        },
        "limitations": [
            (
                "A dHash match is a near-duplicate candidate and may require visual "
                "review; it is conservatively excluded from clean manifests."
            ),
            (
                "An outfit-ID collision alone is not proof of shared source because "
                "EVALUATION3 combines datasets; it is reported separately."
            ),
        ],
    }


def write_audit_outputs(
    output_dir: Path | str,
    audit_rows: Sequence[Mapping[str, object]],
    evidence_rows: Sequence[Mapping[str, object]],
    summary: Mapping[str, object],
) -> dict[str, Path]:
    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)

    paths = {
        "summary": destination / "evaluation3_overlap_summary.json",
        "audit": destination / "evaluation3_overlap_audit.jsonl",
        "full": destination / "evaluation3_full.jsonl",
        "model_clean": destination / "evaluation3_model_clean.jsonl",
        "strict_clean": destination / "evaluation3_strict_clean.jsonl",
        "evidence": destination / "evaluation3_overlap_evidence.csv",
    }
    write_json(summary, paths["summary"])
    write_jsonl(audit_rows, paths["audit"])
    write_jsonl(audit_rows, paths["full"])
    write_jsonl(
        (row for row in audit_rows if row["eligible_for_model_clean"]),
        paths["model_clean"],
    )
    write_jsonl(
        (row for row in audit_rows if row["eligible_for_strict_clean"]),
        paths["strict_clean"],
    )

    evidence_headers = (
        "e3_outfit_id",
        "e3_slot",
        "e3_image",
        "match_type",
        "polyvore_item_id",
        "polyvore_splits",
        "polyvore_source_kit_ids",
        "polyvore_image_source",
        "dhash_hamming_distance",
    )
    with paths["evidence"].open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=evidence_headers, extrasaction="ignore")
        writer.writeheader()
        for row in evidence_rows:
            serialized = dict(row)
            for key in ("polyvore_splits", "polyvore_source_kit_ids"):
                serialized[key] = "|".join(serialized.get(key, []))
            writer.writerow(serialized)
    return paths


def run_overlap_audit(
    *,
    evaluation3_root: Path | str,
    development_split_paths: Mapping[str, Path | str],
    output_dir: Path | str,
    polyvore_image_root: Path | str | None = None,
    polyvore_hf_dataset: str | None = None,
    hf_split_mapping: Mapping[str, str] | None = None,
    annotations_path: Path | str | None = None,
    annotation_sheet: str = "CMT",
    metadata_path: Path | str | None = None,
    metadata_sheet: str = "Num",
    evaluation3_groups: Iterable[str] | None = None,
    model_development_splits: Iterable[str] = DEFAULT_MODEL_DEVELOPMENT_SPLITS,
    near_hamming_threshold: int = DEFAULT_NEAR_HAMMING_THRESHOLD,
    allow_incomplete_image_index: bool = False,
) -> tuple[dict[str, object], dict[str, Path]]:
    if bool(polyvore_image_root) == bool(polyvore_hf_dataset):
        raise ValueError(
            "Choose exactly one Polyvore image provider: --polyvore-image-root "
            "or --polyvore-hf-dataset"
        )

    identity = load_development_identity(development_split_paths)
    item_splits = identity["item_splits"]
    if polyvore_image_root:
        image_fingerprints, provider_report = fingerprint_local_development_images(
            polyvore_image_root, item_splits
        )
    else:
        image_fingerprints, provider_report = fingerprint_huggingface_development_images(
            str(polyvore_hf_dataset), hf_split_mapping or {}, item_splits
        )

    unresolved = sorted(set(item_splits) - set(image_fingerprints))
    if unresolved and not allow_incomplete_image_index:
        raise ValueError(
            f"Polyvore image index is incomplete: {len(unresolved)} of "
            f"{len(item_splits)} scorer item IDs are unresolved. Examples: "
            f"{unresolved[:10]}. Fix the image source or explicitly use "
            "--allow-incomplete-image-index for a provisional audit."
        )

    annotation_tables = []
    if annotations_path:
        annotation_tables.append(
            load_evaluation3_annotations(annotations_path, sheet_name=annotation_sheet)
        )
    if metadata_path:
        annotation_tables.append(
            load_evaluation3_annotations(metadata_path, sheet_name=metadata_sheet)
        )
    annotations = (
        merge_evaluation3_annotations(*annotation_tables)
        if annotation_tables
        else None
    )
    outfits = discover_evaluation3_outfits(
        evaluation3_root,
        annotations=annotations,
        selected_groups=evaluation3_groups,
    )
    if not outfits:
        raise ValueError("No EVALUATION3 outfit folders matched the requested selection")

    development_items = build_development_items(identity, image_fingerprints)
    audit_rows, evidence_rows = audit_overlap(
        evaluation3_root,
        outfits,
        identity,
        development_items,
        model_development_splits=model_development_splits,
        near_hamming_threshold=near_hamming_threshold,
    )
    summary = build_summary(
        audit_rows,
        identity,
        provider_report,
        unresolved_development_item_ids=unresolved,
        model_development_splits=model_development_splits,
        near_hamming_threshold=near_hamming_threshold,
    )
    paths = write_audit_outputs(output_dir, audit_rows, evidence_rows, summary)
    return summary, paths


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Detect EVALUATION3 overlap with scorer Polyvore splits."
    )
    parser.add_argument("--evaluation3-root", required=True, type=Path)
    parser.add_argument(
        "--development-jsonl",
        action="append",
        default=[],
        metavar="NAME=PATH",
        help="Repeat for train/valid/test scorer-ready JSONL files.",
    )
    parser.add_argument("--output-dir", required=True, type=Path)
    provider = parser.add_mutually_exclusive_group(required=True)
    provider.add_argument("--polyvore-image-root", type=Path)
    provider.add_argument(
        "--polyvore-hf-dataset",
        help="For example codewaly/polyvore1000.",
    )
    parser.add_argument(
        "--hf-split",
        action="append",
        default=[],
        metavar="NAME=HF_SPLIT",
        help="Optional mapping if development and Hugging Face split names differ.",
    )
    parser.add_argument("--evaluation3-annotations", type=Path)
    parser.add_argument("--annotation-sheet", default="CMT")
    parser.add_argument(
        "--evaluation3-metadata",
        type=Path,
        help="Attribute_ALL_UBSGsimple workbook used to obtain Group/split.",
    )
    parser.add_argument("--metadata-sheet", default="Num")
    parser.add_argument(
        "--evaluation3-group",
        action="append",
        default=[],
        help="For example A-Test2000. Repeat to select multiple groups.",
    )
    parser.add_argument(
        "--model-development-splits",
        default=",".join(DEFAULT_MODEL_DEVELOPMENT_SPLITS),
        help="Comma-separated splits used for model fitting/selection.",
    )
    parser.add_argument(
        "--near-hamming-threshold",
        default=DEFAULT_NEAR_HAMMING_THRESHOLD,
        type=int,
    )
    parser.add_argument(
        "--allow-incomplete-image-index",
        action="store_true",
        help="Write a provisional audit even when some scorer item images are unresolved.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_argument_parser()
    arguments = parser.parse_args(argv)
    try:
        development_paths = parse_named_paths(
            arguments.development_jsonl, "--development-jsonl"
        )
        hf_split_mapping = (
            {
                name: str(path)
                for name, path in parse_named_paths_without_file_check(
                    arguments.hf_split, "--hf-split"
                ).items()
            }
            if arguments.hf_split
            else {}
        )
        model_splits = {
            value.strip()
            for value in arguments.model_development_splits.split(",")
            if value.strip()
        }
        missing_model_splits = model_splits - set(development_paths)
        if missing_model_splits:
            raise ValueError(
                "model-development splits were not supplied as JSONL: "
                f"{sorted(missing_model_splits)}"
            )
        summary, paths = run_overlap_audit(
            evaluation3_root=arguments.evaluation3_root,
            development_split_paths=development_paths,
            output_dir=arguments.output_dir,
            polyvore_image_root=arguments.polyvore_image_root,
            polyvore_hf_dataset=arguments.polyvore_hf_dataset,
            hf_split_mapping=hf_split_mapping,
            annotations_path=arguments.evaluation3_annotations,
            annotation_sheet=arguments.annotation_sheet,
            metadata_path=arguments.evaluation3_metadata,
            metadata_sheet=arguments.metadata_sheet,
            evaluation3_groups=arguments.evaluation3_group,
            model_development_splits=model_splits,
            near_hamming_threshold=arguments.near_hamming_threshold,
            allow_incomplete_image_index=arguments.allow_incomplete_image_index,
        )
    except (FileNotFoundError, KeyError, RuntimeError, ValueError) as error:
        parser.error(str(error))

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("\nOutputs:")
    for name, path in paths.items():
        print(f"- {name}: {path}")
    return 0


def parse_named_paths_without_file_check(
    values: Sequence[str], option_name: str
) -> dict[str, str]:
    result: dict[str, str] = {}
    for value in values:
        name, separator, target = value.partition("=")
        name = name.strip()
        target = target.strip()
        if not separator or not name or not target:
            raise ValueError(f"{option_name} must use NAME=VALUE, got {value!r}")
        if name in result:
            raise ValueError(f"Duplicate {option_name} name: {name}")
        result[name] = target
    return result


if __name__ == "__main__":
    raise SystemExit(main())
